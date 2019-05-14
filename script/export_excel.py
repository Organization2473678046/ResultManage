# -*- coding: utf-8 -*-
import os
import sys
import psycopg2
import openpyxl
from datetime import datetime
from openpyxl.styles import Font,colors,Alignment



# 写excel
def write_excel(dbname):
    data_list = get_handoutlist_data(dbname)
    f = openpyxl.Workbook()  # 创建工作簿
    sheet1 = f.create_sheet(title="sheet1", index=0)
    # len = 34
    row1 = ["分发单序号", "分发单标题", "编号", "签发", "名称", "用途", "审批编号", "保密责任书编号", "递送方式", "递送方式后的签名", "介质类型", "介质编号", "发出单位",
            "接收单位", "发出单位通讯地址", "接收单位通讯地址", "发出单位邮政编码", "接收单位邮政编码", "经办人", "接收人", "经办人联系电话(座机)", "经办人联系电话(手机)",
            "接收人联系电话(座机)", "接收人联系电话(手机)", "发出日期", "接收日期", "图幅号", "分发单中成果序号", "成果名称", "成果密级", "成果数量", "成果数据量", "格式/介质",
            "备注信息"]
    # 生成第一行
    for i in range(len(row1)):
        sheet1.cell(column=i + 1, row=1).value = row1[i]

    handoutlist_count = 1
    for data_dict in data_list:
        handoutlist = data_dict["handoutlist"]
        fileinfo_list = data_dict["fileinfo_list"]
        fileinfo_count = 1
        for fileinfo in fileinfo_list:
            row = [str(handoutlist_count), handoutlist[0], handoutlist[1], handoutlist[2], handoutlist[3],
                   handoutlist[4], handoutlist[5], handoutlist[6],
                   get_deliverways(handoutlist[7], handoutlist[8], handoutlist[9], handoutlist[0]), handoutlist[11],
                   get_media(handoutlist[12], handoutlist[13], handoutlist[14], handoutlist[15], handoutlist[16]),
                   handoutlist[17], handoutlist[18], handoutlist[19], handoutlist[20], handoutlist[21], handoutlist[22],
                   handoutlist[23], handoutlist[24], handoutlist[25], handoutlist[26], handoutlist[27], handoutlist[28],
                   handoutlist[29], handoutlist[30], handoutlist[31], handoutlist[32], str(fileinfo_count), fileinfo[0],
                   fileinfo[1], fileinfo[2], fileinfo[3], fileinfo[4], fileinfo[5]]
            for item in row:
                if item is None or item == "None":
                    row[row.index(item)] = ""
            if fileinfo_count != 1:
                for i in row:
                    if row.index(i) <= 26:
                        row[row.index(i)] = ""
            sheet1.append(row)
            fileinfo_count += 1
        handoutlist_count += 1

    return f




def get_handoutlist_data(dbname):
    conn = psycopg2.connect(dbname=dbname,
                            user="postgres",
                            password="Lantucx2018",
                            host="localhost",
                            port="5432")
    cursor = conn.cursor()
    # 37个字段
    SELECT_SQL1 = "select title,listnum,signer,name,purpose,auditnum,secrecyagreementnum,selfgetway, postway,networkway,sendtoway,signature,papermedia,cdmedia, diskmedia, networkmedia, othermedia,medianums,sendunit,receiveunit, sendunitaddr,receiveunitaddr, sendunitpostcode,  receiveunitpostcode, handler,receiver,handlerphonenum,handlermobilephonenum,receiverphonenum,receivermobilephonenum, sendouttimec,recievetimec,mapnums, filename,file,uniquenum,id from results_handoutlist order by id"
    cursor.execute(SELECT_SQL1)
    handoutlists = cursor.fetchall()

    data_list = []
    for handoutlist in handoutlists:
        data_dict = {}
        SELECT_SQL2 = "select name,secretlevel,resultnum,datasize,formatormedia,remarks from results_fileinfo where handoutlist_uniquenum='%s' order by id" % handoutlist[35]
        cursor.execute(SELECT_SQL2)
        fileinfo_list = cursor.fetchall()
        data_dict["handoutlist"] = handoutlist
        data_dict["fileinfo_list"]  = fileinfo_list
        data_list.append(data_dict)
        conn.commit()
    print(data_list)
    conn.close()
    return data_list
#
#
# def change_postgresql(dbname):
#     conn = psycopg2.connect(dbname=dbname,
#                             user="postgres",
#                             password="Lantucx2018",
#                             host="localhost",
#                             port="5432")
#     cursor = conn.cursor()
#     sql = "update results_handoutlist set file='{0}' where id={1}".format(hadoutlist_file, handoutlist_id)
#     cursor.execute(sql)
#     conn.commit()
#     conn.close()

def get_deliverways(selfgetway,postway,networkway,sendtoway):
    # 返回清单所选的所有递送方式
    deliverways_list = []
    if selfgetway:
        deliverways_list.append(u"自取")
    if postway:
        deliverways_list.append(u"邮寄")
    if networkway:
        deliverways_list.append(u"网络")
    if sendtoway:
        deliverways_list.append(u"送往")
    return ",".join(deliverways_list)


def get_media(papermedia,cdmedia,diskmedia,networkmedia,othermedia):
    # 返回成果所选的所有介质
    media_list = []
    if papermedia:
        media_list.append(u"纸质")
    if cdmedia:
        media_list.append(u"光盘")
    if diskmedia:
        media_list.append(u"硬盘")
    if networkmedia:
        media_list.append(u"网络")
    if othermedia:
        media_list.append(u"其他")

    return ",".join(media_list)


if __name__ == '__main__':
    # get_handoutlist_fileinfo_data("resmanagev0.3")
    # 写入Excel
    write_excel("resmanagev0.3","/opt/rh/httpd24/root/var/www/html/ResultManage/media/handoutlist_excels")
    print('写入成功')

